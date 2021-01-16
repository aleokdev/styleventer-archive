print("Importing modules...")
from openpyxl import load_workbook
import time,sys

print("Creating bindings...")
nameBindings ={}
f = open(sys.argv[1] + "/rpgBind.txt", "r")
for nBind in f.read().split("\n"):
	if nBind == "":
		continue
		
	nameBindings[nBind.split(":")[0]] = (nBind.split(":")[1].split(",")[0], nBind.split(":")[1].split(",")[1])
	print(nBind.split(":")[0] + ": "+ nameBindings[nBind.split(":")[0]][0])
f.close()

weaponBindings ={
	"System Name":"* ",
	"Type":"*t",
	"Attack":"*a",
	"Agility":"*l",
	"One/Double Handed":"*d",
	"Attributes":"*-->",
	"Loot Type:Rarity":"*r"
}

attributeBindings ={
	"Name (M)":"^ ",
	"Name (F)":"^f",
	"Type":"^t",
	"Classification":"^c",
	"Attack Mod":"^a",
	"Agility Mod":"^g",
	"Probability":"^p"
}

biomeBindings ={
	"System Name":"& ",
	"Connections":"&->",
	"Emoji":"&e",
	"Properties":"&p"
}

locationBindings ={
	"System Name":"$ ",
	"Connections":"$->",
	"Loot Type":"$l",
	"Extra Attributes":"$a"
}

enemyBindings ={
	"System Name":"! ",
	"Gender":"!g",
	"HP":"!h",
	"Atk Multiplier":"!m",
	"Weapons":"!w",
	"Attributes":"!a",
	"Found in:Rarity":"!l",
	"XP dropped":"!x",
	"Special Attacks":"!s"
}

sheetBindings ={
"Weapons":weaponBindings,
"Attributes":attributeBindings,
"Biomes":biomeBindings,
"Locations":locationBindings,
"Enemies":enemyBindings
}

columnLetters = "ABCDEFGHIJKMNLOPQRSTUVWXYZ"
print("Loading worksheet...")
wb = load_workbook("rpgData.xlsx")
data = ""
print("Parsing data...")
for sheetname in wb.sheetnames:
	s = wb[sheetname]
	for row in s.iter_rows(min_row=2, max_col=len(sheetBindings[sheetname]), max_row=50):
		if row[0].value == None:
			break
		i = 0
		for cell in row:
			try:
				key = sheetBindings[sheetname][s[columnLetters[i]+"1"].value]
			except KeyError:
				print("'{}1' doesn't exist on sheet!".format(columnLetters[i]))
			if s[columnLetters[i]+"1"].value == "System Name":
				toAdd = nameBindings[cell.value]
				# Add system name, name and gender
				data = data + ((key.rstrip()) * 2) + cell.value + "\n" + key + toAdd[0] + "\n" + key.rstrip() + "g" + toAdd[1].upper() + "\n"
				i = i + 1
				continue
			
			value = key + ("\n" + key).join(str(cell.value).split(","))
			
			if value != key + "None":
				data = data + value + "\n"
			i = i + 1
		
data = data+"\neof"
				
print("Transferring data...")
f = open(sys.argv[1] + "/rpgData.txt","w", encoding="utf-8")
f.write(data)
f.close()

print("Used {} lines.".format(data.count("\n") + 1))
print("Finished in {}s.".format(round(time.process_time(), 2)))